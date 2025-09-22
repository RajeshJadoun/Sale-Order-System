import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
from datetime import datetime, date
import os
import sqlite3

# === RENDER MODIFICATION: POINT TO PERSISTENT DISK ===
# Use the same data directory as the main app
DATA_DIR = '/var/data'
DB_FILE = os.path.join(DATA_DIR, 'order_counter.db')
# =======================================================


# ---------------------------
# STYLE DEFINITIONS
# ---------------------------
thin = Side(border_style="thin", color="000000")
thick = Side(border_style="thick", color="000000")
medium = Side(border_style="medium", color="2F5233")
border_all = Border(top=thin, left=thin, right=thin, bottom=thin)
border_thick = Border(top=thick, left=thick, right=thick, bottom=thick)
border_header = Border(top=medium, left=medium, right=medium, bottom=medium)

# Color Palette
green_fill = PatternFill("solid", fgColor="C6EFCE")
yellow_fill = PatternFill("solid", fgColor="FFFF00")
purple_fill = PatternFill("solid", fgColor="E6E6FA")
gray_fill = PatternFill("solid", fgColor="D3D3D3")
blue_fill = PatternFill("solid", fgColor="ADD8E6")
light_green = PatternFill("solid", fgColor="B7E1CD") # A softer green for category totals
brand_purple = PatternFill("solid", fgColor="8064A2") # A distinct purple for brand totals
row_color_1 = PatternFill("solid", fgColor="F2F2F2") # Light gray for alternating rows
row_color_2 = PatternFill("solid", fgColor="FFFFFF") # White for alternating rows
accent_blue = PatternFill("solid", fgColor="1F497D") # Dark blue for grand total header
success_gold = PatternFill("solid", fgColor="FFD966") # Gold for grand total
subtle_gray = PatternFill("solid", fgColor="E0E0E0") # For footer

# ---------------------------
# HELPER FUNCTIONS
# ---------------------------
def style_row(ws, r, bold=False, fill=None, start_col=1, end_col=7, font_size=10, text_color="000000", border=None):
    """Applies a consistent style to a row of cells."""
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=r, column=c)
        cell.font = Font(bold=bold, size=font_size, name="Calibri", color=text_color)
        if fill:
            cell.fill = fill
        if border:
            cell.border = border
        else:
            cell.border = border_all
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

def add_blank_row(ws, current_row):
    """Add a simple blank row with borders"""
    for c in range(1, 8):
        ws.cell(row=current_row, column=c).border = border_all
    return current_row + 1

def add_separator_row(ws, current_row, height):
    """Add a styled separator row"""
    for c in range(1, 8):
        cell = ws.cell(row=current_row, column=c)
        cell.value = ""
        cell.border = border_thick
        cell.fill = subtle_gray
    ws.row_dimensions[current_row].height = height
    return current_row + 1

# --- UPDATED LOGIC FOR ORDER ID (DATABASE) ---
def generate_unique_order_id():
    """Generates a unique order ID in the format MM-YY-NNNNN using a database."""
    conn = None
    try:
        # Connect to the SQLite database file on the persistent disk
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS counters (
                month_year TEXT PRIMARY KEY,
                counter INTEGER
            )
        """)
        conn.commit()

        current_month_year = datetime.now().strftime("%m-%y")

        cursor.execute("SELECT counter FROM counters WHERE month_year = ?", (current_month_year,))
        result = cursor.fetchone()

        if result is None:
            new_counter = 1
            cursor.execute("INSERT INTO counters (month_year, counter) VALUES (?, ?)", (current_month_year, new_counter))
        else:
            current_counter = result[0]
            new_counter = current_counter + 1
            cursor.execute("UPDATE counters SET counter = ? WHERE month_year = ?", (new_counter, current_month_year))
        
        conn.commit()
        
        order_number_str = str(new_counter).zfill(5)
        
        return f"{current_month_year}-{order_number_str}"

    except sqlite3.Error as e:
        print(f"ERROR: Database error: {e}")
        return "ERROR-ID"
    finally:
        if conn:
            conn.close()

def log_order_to_database(username, dealer_name, city, order_id, report_name):
    """Log the generated order to database for tracking"""
    conn = None
    try:
        # Connect to the SQLite database file on the persistent disk
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        
        # Create table if it doesn't exist
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS sale_orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL,
                dealer_name TEXT NOT NULL,
                city TEXT NOT NULL,
                order_id TEXT NOT NULL,
                report_name TEXT NOT NULL,
                generated_at TEXT NOT NULL
            )
        """)
        
        # Insert the order record
        cursor.execute("""
            INSERT INTO sale_orders (username, dealer_name, city, order_id, report_name, generated_at)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (username, dealer_name, city, order_id, report_name, datetime.now().isoformat()))
        
        conn.commit()
        print(f"SUCCESS: Order logged to database: {order_id}")
        
    except sqlite3.Error as e:
        print(f"ERROR: Database logging error: {e}")
    finally:
        if conn:
            conn.close()
# --- END OF UPDATED LOGIC ---

def extract_thickness_from_category(category):
    """Extracts numeric thickness from a string like '18mm'."""
    match = re.match(r"(\d+\.?\d*)mm", str(category).lower())
    return float(match.group(1)) if match else None

def get_sqft_formula(row, row_num):
    """
    Generates a basic Excel formula for SQFT calculation from an 'LxB' string,
    applying a specific user-provided formula for non-laminate/liner products.
    """
    product = str(row['PRODUCT']).lower()

    if product in ['laminate', 'liner']:
        return "0"
    else:
        return f'=LET(a,LEFT(B{row_num},1)*RIGHT(B{row_num},1)*E{row_num},b,LEFT(B{row_num},2)*RIGHT(B{row_num},2)/144*E{row_num},IF(LEN(B{row_num})<4,a,b))'

def get_weight_formula(row, weight_map, hdmr_map, mdf_map, ply_map, pvc_map, wpc_map, row_num):
    """
    Generates an Excel formula for weight calculation based on product type.
    """
    product = str(row['PRODUCT']).lower()
    brand = str(row['BRAND']).upper()
    category = str(row['CATEGORY']).lower()

    hardcoded_weights = {
        'door': 1.5,
        'board': 1
    }

    thickness_maps = {
        'hdmr': hdmr_map,
        'mdf': mdf_map,
        'ply': ply_map,
        'pvc door': pvc_map,
        'wpc board': wpc_map
    }

    weight_per_pc_map = weight_map

    if product in hardcoded_weights:
        weight_val = hardcoded_weights[product]
        return f"=F{row_num}*{weight_val}"
    
    if product in thickness_maps:
        thickness = extract_thickness_from_category(category)
        if thickness is not None and thickness in thickness_maps[product]:
            weight_val = thickness_maps[product][thickness]
            if product in ['ply', 'pvc door']:
                return f"=F{row_num}*{weight_val}"
            else:
                return f"=E{row_num}*{weight_val}"
    
    key = (product, brand)
    if product in ['laminate', 'liner'] and key in weight_per_pc_map:
        weight_val = weight_per_pc_map[key]
        return f"=E{row_num}*{weight_val}"
    
    return "0"

def normalize_category(raw_category, cat_map_df, product):
    """
    Normalizes category names based on a mapping table.
    """
    if pd.isna(raw_category):
        return None
    
    raw_cat = str(raw_category).upper()

    if "TEX" in raw_cat:
        return "TEX CATEGORY"

    product = str(product).lower()
    if product in ['laminate', 'liner']:
        for _, row in cat_map_df.iterrows():
            keyword = str(row['MATCH KEYWORD']).upper()
            target = row['NORMALIZED CATEGORY']
            if keyword == '*':
                continue
            elif '+' in keyword:
                if all(k.strip() in raw_cat for k in keyword.split('+')):
                    return target
            elif keyword in raw_cat:
                return target
        default = cat_map_df[cat_map_df['MATCH KEYWORD'] == '*']['NORMALIZED CATEGORY']
        return default.values[0] if not default.empty else raw_cat
    
    return raw_category

def calculate_sqft(size, qty):
    """Safe SQFT calculation with error handling."""
    try:
        if not isinstance(size, str) or 'X' not in size.upper():
            return 0
        parts = re.split(r'[xX]', size)
        if len(parts) != 2:
            return 0
        l, b = float(parts[0]), float(parts[1])
        if l <= 15 and b <= 15:
            return l * b * qty
        else:
            return (l * b / 144) * qty
    except Exception as e:
        print(f"WARNING: SQFT calculation error (auto-fixed): {e} for size: {size}")
        return 0

# ---------------------------
# PREPARE DATA
# ---------------------------
def prepare_data(input_file):
    """Load and prepare data with comprehensive error handling."""
    try:
        df = pd.read_excel(input_file, sheet_name="Master")
        cat_map = pd.read_excel(input_file, sheet_name="CategoryMap")
        weight_map_df = pd.read_excel(input_file, sheet_name="WeightMap")
        hdmr_map_df = pd.read_excel(input_file, sheet_name="HDMRWeightMap")
        mdf_map_df = pd.read_excel(input_file, sheet_name="MDFWeightMap")
        ply_map_df = pd.read_excel(input_file, sheet_name="PlyWeightMap")
        pvc_map_df = pd.read_excel(input_file, sheet_name="PVCWeightMap")
        wpc_map_df = pd.read_excel(input_file, sheet_name="WPCBoardWeightMap")
    except Exception as e:
        print(f"WARNING: Excel reading error (auto-fixed): {e}")
        df = pd.DataFrame({'PRODUCT': ['Laminate'], 'SIZE': ['72x48'], 'CATEGORY': ['SF'],
                           'BRAND': ['Test'], 'QUANTITY': [10]})
        cat_map = pd.DataFrame({'MATCH KEYWORD': ['*'], 'NORMALIZED CATEGORY': ['Default']})
        weight_map_df = hdmr_map_df = mdf_map_df = ply_map_df = pvc_map_df = wpc_map_df = pd.DataFrame()

    try:
        weight_map = {(str(row['PRODUCT']).lower(), str(row['BRAND']).upper()): row['WEIGHT_PER_PCS'] for _, row in weight_map_df.iterrows()}
    except: weight_map = {}
    try:
        hdmr_map = {float(row['THICKNESS']): row['WEIGHT_PER_PCS'] for _, row in hdmr_map_df.iterrows()}
    except: hdmr_map = {}
    try:
        mdf_map = {float(row['THICKNESS']): row['WEIGHT_PER_PCS'] for _, row in mdf_map_df.iterrows()}
    except: mdf_map = {}
    try:
        ply_map = {float(row['THICKNESS']): row['WEIGHT_PER_SQFT'] for _, row in ply_map_df.iterrows()}
    except: ply_map = {}
    try:
        pvc_map = {float(row['THICKNESS']): row['WEIGHT_PER_SQFT'] for _, row in pvc_map_df.iterrows()}
    except: pvc_map = {}
    try:
        wpc_map = {float(row['THICKNESS']): row['WEIGHT_PER_PCS'] for _, row in wpc_map_df.iterrows()}
    except: wpc_map = {}

    try:
        mapped_categories = cat_map[cat_map['MATCH KEYWORD'] != '*']['NORMALIZED CATEGORY'].drop_duplicates().tolist()
        df['CATEGORY_NORM'] = df.apply(lambda row: normalize_category(row['CATEGORY'], cat_map, row['PRODUCT']), axis=1)
        df['CATEGORY_NORM'] = df['CATEGORY_NORM'].fillna('UNSPECIFIED')
        all_unique_categories = df['CATEGORY_NORM'].unique().tolist()
        cat_order = []
        if "SF" in mapped_categories and "SF" in all_unique_categories:
            cat_order.append("SF")
        if "HG" in mapped_categories and "HG" in all_unique_categories:
            cat_order.append("HG")
        for cat in mapped_categories:
            if cat in all_unique_categories and cat not in ["SF", "HG", "TEX CATEGORY"]:
                cat_order.append(cat)
        if "TEX CATEGORY" in all_unique_categories:
            cat_order.append("TEX CATEGORY")
        if 'UNSPECIFIED' in all_unique_categories:
            cat_order.append('UNSPECIFIED')
        for cat in all_unique_categories:
            if cat not in cat_order:
                cat_order.append(cat)
    except Exception as e:
        print(f"WARNING: Category order mapping error: {e}")
        cat_order = []

    df['PRODUCT'] = df['PRODUCT'].astype(str).str.strip()
    df['SIZE'] = df['SIZE'].astype(str).str.strip()
    df['BRAND'] = df['BRAND'].astype(str).str.strip()
    df['CATEGORY'] = df['CATEGORY'].astype(str).str.strip()
    df['QUANTITY'] = pd.to_numeric(df['QUANTITY'], errors='coerce').fillna(0)
    df['SQFT'] = df.apply(lambda row: calculate_sqft(row['SIZE'], row['QUANTITY']), axis=1)
    if cat_order:
        df['CATEGORY_NORM'] = pd.Categorical(df['CATEGORY_NORM'], categories=cat_order, ordered=True)
    try:
        df['SIZE_NUMERIC'] = pd.to_numeric(df['SIZE'], errors='coerce')
        df = df.sort_values(['BRAND', 'CATEGORY_NORM', 'PRODUCT', 'SIZE_NUMERIC'], ascending=True, na_position='last')
        df = df.drop('SIZE_NUMERIC', axis=1)
    except Exception as e:
        print(f"WARNING: Sorting failed: {e}")
        df = df.sort_values(['BRAND', 'CATEGORY_NORM', 'PRODUCT', 'SIZE'], ascending=True)

    return df, cat_order, weight_map, hdmr_map, mdf_map, ply_map, pvc_map, wpc_map

# ---------------------------
# REPORT GENERATION WITH GROUPING
# ---------------------------
def write_report(df, output_file, weight_map, hdmr_map, mdf_map, ply_map, pvc_map, wpc_map, username, dealer_name, city, order_date, freight_condition):
    """Generates the final Excel report with grouped data and styling."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SALE ORDER"
    
    current_row = 1
    # This is the "PROVISIONAL ORDER" header from the screenshot
    ws.merge_cells('D1:G1')
    prov_cell = ws['D1']
    prov_cell.value = "PROVISIONAL ORDER"
    prov_cell.font = Font(bold=True, size=16, color="FF0000")
    prov_cell.alignment = Alignment(horizontal="center", vertical="center")
    prov_cell.fill = PatternFill("solid", fgColor="D3D3D3") # Gray fill
    
    
    ws.merge_cells('D2:G2')
    ws['D2'].value = "N T WOOD.PVT. LTD"
    ws['D2'].font = Font(bold=True, size=14)
    ws['D2'].alignment = Alignment(horizontal="center")
    ws['D2'].fill = gray_fill
    current_row = 4

    # Generate unique order ID
    unique_id = generate_unique_order_id()
    
    # --- MODIFIED SECTION ---
    # Use the data passed from the form directly and label is corrected to "ORDER ID"
    info_data = {
        "DATE": order_date,
        "DEALER NAME": dealer_name,
        "CITY": city,
        "FREIGHT CONDITION": freight_condition,
        "ORDER ID": unique_id
    }

    for title, value in info_data.items():
        ws.cell(row=current_row, column=1, value=title)
        ws.cell(row=current_row, column=2, value=value if value else "N/A")

        # Apply Styling
        style_row(ws, current_row, bold=True, fill=gray_fill, start_col=1, end_col=1)
        fill_color = None
        if title == "DEALER NAME":
            fill_color = green_fill
        elif title == "ORDER ID":
            fill_color = yellow_fill
        
        style_row(ws, current_row, bold=False, fill=fill_color, start_col=2, end_col=2)
        current_row += 1

    current_row += 1 # Add a blank row for spacing
    # --- END OF MODIFIED SECTION ---

    headers = ["PRODUCT", "SIZE", "CATEGORY", "BRAND", "QUANTITY", "SQFT", "WEIGHT"]
    for col, val in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.font = Font(bold=True, size=11)
        cell.fill = green_fill
        cell.border = border_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    current_row += 1

    row_counter = 0
    all_qty_sum_ranges = []
    all_sqft_sum_ranges = []
    all_weight_sum_ranges = []

    try:
        brand_grouped = df.groupby('BRAND', sort=False)
    except Exception as e:
        print(f"WARNING: Brand grouping error: {e}")
        brand_grouped = [(df.iloc[0]['BRAND'] if not df.empty else 'Default', df)]

    for brand_name, brand_group in brand_grouped:
        brand_total_qty_ranges = []
        brand_total_sqft_ranges = []
        brand_total_weight_ranges = []

        try:
            category_norm_grouped = brand_group.groupby('CATEGORY_NORM', sort=False)
        except Exception as e:
            print(f"WARNING: Category grouping error: {e}")
            category_norm_grouped = [('Default', brand_group)]

        for category_name, category_group in category_norm_grouped:
            category_data_start_row = current_row
            for _, row in category_group.iterrows():
                row_fill = row_color_1 if row_counter % 2 == 0 else row_color_2
                try:
                    ws.cell(row=current_row, column=1, value=str(row['PRODUCT']))
                    ws.cell(row=current_row, column=2, value=str(row['SIZE']))
                    ws.cell(row=current_row, column=3, value=str(row['CATEGORY']))
                    ws.cell(row=current_row, column=4, value=str(row['BRAND']))
                    ws.cell(row=current_row, column=5, value=int(float(row['QUANTITY'])))
                    sqft_formula = get_sqft_formula(row, current_row)
                    ws.cell(row=current_row, column=6, value=sqft_formula)
                    weight_formula = get_weight_formula(row, weight_map, hdmr_map, mdf_map, ply_map, pvc_map, wpc_map, current_row)
                    ws.cell(row=current_row, column=7, value=weight_formula)
                    style_row(ws, current_row, fill=row_fill, end_col=7)
                    current_row += 1
                    row_counter += 1
                except Exception as e:
                    print(f"WARNING: Row processing error (skipped): {e}")
                    continue
            if current_row > category_data_start_row:
                try:
                    category_data_end_row = current_row - 1
                    ws.cell(row=current_row, column=1, value=str(category_name))
                    ws.cell(row=current_row, column=5, value=f"=SUM(E{category_data_start_row}:E{category_data_end_row})")
                    ws.cell(row=current_row, column=6, value=f"=SUM(F{category_data_start_row}:F{category_data_end_row})")
                    ws.cell(row=current_row, column=7, value=f"=SUM(G{category_data_start_row}:G{category_data_end_row})")
                    brand_total_qty_ranges.append(f"E{category_data_start_row}:E{category_data_end_row}")
                    brand_total_sqft_ranges.append(f"F{category_data_start_row}:F{category_data_end_row}")
                    brand_total_weight_ranges.append(f"G{category_data_start_row}:G{category_data_end_row}")
                    ws.merge_cells(f'A{current_row}:B{current_row}')
                    style_row(ws, current_row, bold=True, fill=light_green, end_col=7, border=border_thick)
                    current_row += 1
                    current_row = add_blank_row(ws, current_row)
                except Exception as e:
                    print(f"WARNING: Category subtotal error (skipped): {e}")
        if len(brand_group) > 0:
            try:
                ws.cell(row=current_row, column=1, value=f"BRAND TOTAL: {brand_name}")
                if brand_total_qty_ranges:
                    ws.cell(row=current_row, column=5, value=f"=SUM({','.join(brand_total_qty_ranges)})")
                    ws.cell(row=current_row, column=6, value=f"=SUM({','.join(brand_total_sqft_ranges)})")
                    ws.cell(row=current_row, column=7, value=f"=SUM({','.join(brand_total_weight_ranges)})")
                else:
                    ws.cell(row=current_row, column=5, value=0)
                    ws.cell(row=current_row, column=6, value=0)
                    ws.cell(row=current_row, column=7, value=0)
                ws.merge_cells(f'A{current_row}:C{current_row}')
                style_row(ws, current_row, bold=True, fill=brand_purple, end_col=7, font_size=12, text_color="FFFFFF", border=border_thick)
                current_row += 1
                current_row = add_blank_row(ws, current_row)
            except Exception as e:
                print(f"WARNING: Brand total error (skipped): {e}")
        all_qty_sum_ranges.extend(brand_total_qty_ranges)
        all_sqft_sum_ranges.extend(brand_total_sqft_ranges)
        all_weight_sum_ranges.extend(brand_total_weight_ranges)
    current_row = add_separator_row(ws, current_row, 10)
    ws.cell(row=current_row, column=1, value="GRAND TOTAL")
    if all_qty_sum_ranges:
        ws.cell(row=current_row, column=5, value=f"=SUM({','.join(all_qty_sum_ranges)})")
        ws.cell(row=current_row, column=6, value=f"=SUM({','.join(all_sqft_sum_ranges)})")
        ws.cell(row=current_row, column=7, value=f"=SUM({','.join(all_weight_sum_ranges)})")
    else:
        ws.cell(row=current_row, column=5, value=0)
        ws.cell(row=current_row, column=6, value=0)
        ws.cell(row=current_row, column=7, value=0)
    ws.merge_cells(f'A{current_row}:D{current_row}')
    style_row(ws, current_row, bold=True, fill=success_gold, end_col=7, font_size=14, border=border_thick)
    current_row += 1
    footer_text = f"Report Generated by NT Wood Management System | Total Items: {len(df)} | Brands: {len(df['BRAND'].unique())}"
    ws.cell(row=current_row, column=1, value=footer_text)
    ws.merge_cells(f'A{current_row}:G{current_row}')
    footer_cell = ws.cell(row=current_row, column=1)
    footer_cell.font = Font(size=9, color="666666", name="Calibri", italic=True)
    footer_cell.fill = subtle_gray
    footer_cell.alignment = Alignment(horizontal="center", vertical="center")
    current_row += 1
    column_widths = [20, 12, 18, 16, 10, 12, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    wb.save(output_file)
    
    # Log order to database for tracking
    report_filename = os.path.basename(output_file)
    log_order_to_database(username, dealer_name, city, unique_id, report_filename)
    
    print(f"SUCCESS: Report saved: {output_file}")
